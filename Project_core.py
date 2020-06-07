import openpyxl
def search_cell (condition, Sheet):
    for row in Sheet.iter_rows():
            for cell in row:
                if (cell.value == condition):
                    return (cell)
    return (None)
data = openpyxl.load_workbook('data.xlsx')
Sheet = data.active
print ("Wellcome to your post-quarantine activity manager!")
while (True):
    print("You can 'add' a new activity, 'edit' an existing activity, 'delete' an activity, 'search' for anything specific or 'exit'")
    print ("What would you like to do?")
    ans = str(input())
    if (ans == "exit"):
        data.save('data.xlsx')
        print("Closing down")
        break
    elif (ans == "add"):
        print ("Do you want to create a new category?")
        print ("Type 'yes' to create a category or 'no' if you dont want to do that")
        ans = str(input())
        if (ans == 'yes'):
            print ("Please type in the name of the category")
            flag = True
            ans = str(input())
            for cell in Sheet['1']:
                if (cell.value == ans):
                    flag = False
                    print ("This category already exists")
                    break
            if (flag):
                Sheet[(chr(len(Sheet['1'])+65)) + '1'] = ans
        else:
            print ("Which category do you want to add an activity to?")
            print ("Here are your possible categories:")
            for cell in Sheet['1']:
                print (cell.value)
            ans = str(input())
            cell0 = search_cell(ans, Sheet)
            if (cell0 == None):
                print ("There is no such category")
            else:
                print ("What do you want to do after quarantine?")
                ans = str(input())
                none_count = None
                flag = False
                for cell in Sheet[chr(cell0.column + 64)]:
                    if (cell.value == None):
                        if (none_count == None):
                            none_count = cell
                    if (cell.value == ans):
                        flag = True
                        break
                if (flag):
                    print("This activity is already in the database")
                else:
                    if (none_count != None):
                        Sheet[none_count.coordinate] = ans
                    else:
                        Sheet[(chr(cell0.column + 64))+ str(len(Sheet[chr(cell0.column + 64)])+1)] = ans
                    print("Data saved")
    elif (ans == "edit"):
        print("Please type in the name of the activity that you want to edit and what you want to change it to (with a space bar in between)")
        ans, n_ans = map(str, input().split())
        cell = search_cell (ans, Sheet)
        if (cell == None):
            print ("There is no such activity")
        else:
            Sheet[cell.coordinate] = n_ans
    elif (ans == "delete"):
        print("Please type in the name of the activity that you want to delete")
        ans = str(input())
        flag = True
        cell = search_cell (ans, Sheet)
        if (cell == None):
            print ("There is no such activity")
        else:
            Sheet[cell.coordinate] = None
    elif (ans == "search"):
        print ("Please type in the first few letters of the activity, that you are searching for")
        ans = str(input())
        print ("Here are all activities that look like what you are searching for:")
        for row in Sheet.iter_rows():
            if (row != Sheet['1']):
                for cell in row:
                    if (ans in str(Sheet[cell.coordinate].value)):
                        print(cell.value)
    else:     
        print("It seems your input was incorrect, please try again")
